"""Import the シリアル照合表 Excel into the SQLite database.

Reads the "全機種　シリアル照合表" sheet (or any sheet with the same layout).
Each row with a serial number becomes a unit. Rows with an empty 出荷日
become in-stock units (status='在庫'); rows with a value in 出荷日 become
shipped units (status='出荷済') with a sales record.

Usage:
    python import_excel.py "path/to/シリアル管理表.xlsx"
"""
from __future__ import annotations

import re
import sys
from datetime import datetime
from pathlib import Path

import openpyxl

from config import DB_PATH
from db import connect, init_db


SHEET_CANDIDATES = ["全機種　シリアル照合表", "全機種 シリアル照合表"]
HEADER_ROW = 5
DATA_START_ROW = 6

COL = {
    "seq":         2,   # B 通番
    "lot":         3,   # C ロット
    "serial":      4,   # D シリアルNo.
    "model":       5,   # E 機種
    "power":       6,   # F 出力
    "controller":  7,   # G コントローラ
    "mfg_date":    8,   # H 製造年月日表記
    "remark":      9,   # I 備考
    "arrival":    10,   # J 入荷日
    "ship":       11,   # K 出荷日
    "buyer":      12,   # L 購入者
    "address":    13,   # M 住所
    "product":    14,   # N 販売商品名
    "info":       16,   # P 情報
    "price":      18,   # R 決算時単価計算（税抜き価格）
}


def _pick_sheet(wb):
    for name in SHEET_CANDIDATES:
        if name in wb.sheetnames:
            return wb[name]
    for name in wb.sheetnames:
        if "シリアル照合表" in name or "全機種" in name:
            return wb[name]
    raise ValueError(
        f"対象シートが見つかりません。期待するシート名: {SHEET_CANDIDATES} / "
        f"このファイルのシート: {wb.sheetnames}"
    )


def _unmerge_in_memory(ws) -> None:
    """Propagate the top-left value of every merged range to all cells in the
    range. Unmerges in-memory so .cell(row, col).value works uniformly for
    every row covered by the merge."""
    ranges = [str(r) for r in ws.merged_cells.ranges]
    for ref in ranges:
        merged_range = ws[ref]
        top_left = merged_range[0][0]
        value = top_left.value
        ws.unmerge_cells(ref)
        for row in merged_range:
            for cell in row:
                ws.cell(row=cell.row, column=cell.column).value = value


def _s(v):
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    return s if s else None


_DATE_RE = re.compile(r"(\d{4})[./年-](\d{1,2})[./月-](\d{1,2})")


def _date_str(v):
    """Return YYYY-MM-DD if v looks like a date (datetime or text), else None.

    Accepts strings like '2024.12.20', '2024/12/20', '2024-12-20',
    '2024年12月20日', or text containing a date prefix.
    """
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    if not s:
        return None
    m = _DATE_RE.search(s)
    if not m:
        return None
    y, mo, d = m.groups()
    try:
        return datetime(int(y), int(mo), int(d)).strftime("%Y-%m-%d")
    except ValueError:
        return None


def _i(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return int(v)
    s = str(v).strip().replace(",", "").replace("円", "").replace("¥", "")
    try:
        return int(float(s))
    except ValueError:
        return None


def _join_memo(*labeled):
    """labeled: iterable of (label, value); skip empty values."""
    parts = []
    for label, v in labeled:
        s = _s(v)
        if s:
            parts.append(f"{label}: {s}" if label else s)
    return "\n".join(parts) if parts else None


def _is_shipped(ship_val) -> bool:
    """出荷日 column populated in any form => shipped."""
    if ship_val is None:
        return False
    if isinstance(ship_val, datetime):
        return True
    s = str(ship_val).strip()
    if not s or s in ("-", "－", "—"):
        return False
    return True


def import_excel(xlsx_path: Path) -> tuple[int, int]:
    """Returns (imported_count, skipped_count)."""
    init_db()
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = _pick_sheet(wb)
    _unmerge_in_memory(ws)

    imported = 0
    skipped = 0

    with connect(DB_PATH) as conn:
        for r in range(DATA_START_ROW, ws.max_row + 1):
            def get(key):
                return ws.cell(row=r, column=COL[key]).value

            serial = _s(get("serial"))
            model = _s(get("model"))

            if not serial:
                skipped += 1
                continue

            if not model:
                model = "(機種不明)"

            ship_val = get("ship")
            shipped = _is_shipped(ship_val)
            status = "出荷済" if shipped else "在庫"

            unit_memo = _join_memo(
                ("通番", get("seq")),
                ("出力", get("power")),
                ("コントローラ", get("controller")),
                ("備考", get("remark")),
                ("情報", get("info")),
            )

            cur = conn.execute(
                """INSERT INTO units(serial_no, model, mfg_date, status, memo)
                   VALUES (?, ?, ?, ?, ?)""",
                (serial, model, _s(get("mfg_date")), status, unit_memo),
            )
            unit_id = cur.lastrowid

            # Purchase (入荷) record — lot column carries vendor info
            arr_date = _date_str(get("arrival"))
            price = _i(get("price"))
            vendor = _s(get("lot"))
            if arr_date or price or vendor:
                conn.execute(
                    """INSERT INTO purchases(unit_id, purchase_date, vendor_name, amount)
                       VALUES (?, ?, ?, ?)""",
                    (unit_id, arr_date, vendor, price),
                )

            # Sale (出荷) record — only if shipped
            if shipped:
                sale_date = _date_str(ship_val)
                sale_memo_parts = []
                if sale_date is None and ship_val is not None:
                    sale_memo_parts.append(f"出荷日(原文): {ship_val}")
                product = _s(get("product"))
                if product and product != "*":
                    sale_memo_parts.append(f"販売商品名: {product}")
                sale_memo = "\n".join(sale_memo_parts) if sale_memo_parts else None

                conn.execute(
                    """INSERT INTO sales(
                        unit_id, sale_date,
                        customer_name, address,
                        memo
                    ) VALUES (?, ?, ?, ?, ?)""",
                    (
                        unit_id,
                        sale_date,
                        _s(get("buyer")),
                        _s(get("address")),
                        sale_memo,
                    ),
                )

            imported += 1

    return imported, skipped


def main():
    if len(sys.argv) < 2:
        print("Usage: python import_excel.py <path-to-xlsx>")
        sys.exit(1)
    xlsx = Path(sys.argv[1])
    if not xlsx.exists():
        print(f"File not found: {xlsx}")
        sys.exit(1)
    imp, skp = import_excel(xlsx)
    print(f"Imported: {imp} rows / Skipped: {skp} rows")
    print(f"Database: {DB_PATH}")


if __name__ == "__main__":
    main()

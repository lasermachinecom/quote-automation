"""End-to-end test for import_excel.import_excel().

Builds a tiny in-memory workbook matching the シリアル照合表 layout and
verifies that import yields the expected unit / purchase / sales rows.
"""
from __future__ import annotations

import pytest

openpyxl = pytest.importorskip("openpyxl")

import db as db_mod  # noqa: E402
import import_excel as ie  # noqa: E402
from openpyxl.styles import PatternFill  # noqa: E402


YELLOW = PatternFill(patternType="solid", fgColor="FFFFFF00")  # 在庫
GREEN = PatternFill(patternType="solid", fgColor="FF92D050")   # 資産


def _build_workbook(rows):
    """rows: list of dicts with keys serial/model/arrival/ship/buyer/price/fill"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "全機種　シリアル照合表"

    # Header row at row 5 — values don't matter; only the layout matters.
    headers = {
        2: "通番", 3: "ロット", 4: "シリアル", 5: "機種", 6: "出力",
        7: "コントローラ", 8: "製造年月日", 9: "備考", 10: "入荷日",
        11: "出荷日", 12: "購入者", 13: "住所", 14: "販売商品名",
        16: "情報", 18: "決算時単価",
    }
    for col, label in headers.items():
        ws.cell(row=ie.HEADER_ROW, column=col, value=label)

    for i, row in enumerate(rows):
        r = ie.DATA_START_ROW + i
        ws.cell(row=r, column=ie.COL["serial"], value=row.get("serial"))
        ws.cell(row=r, column=ie.COL["model"], value=row.get("model"))
        ws.cell(row=r, column=ie.COL["arrival"], value=row.get("arrival"))
        ws.cell(row=r, column=ie.COL["ship"], value=row.get("ship"))
        ws.cell(row=r, column=ie.COL["buyer"], value=row.get("buyer"))
        ws.cell(row=r, column=ie.COL["price"], value=row.get("price"))
        ws.cell(row=r, column=ie.COL["lot"], value=row.get("lot"))
        ws.cell(row=r, column=ie.COL["address"], value=row.get("address"))

        fill = row.get("fill")
        if fill is not None:
            ws.cell(row=r, column=ie.COL["serial"]).fill = fill

    return wb


@pytest.fixture
def patched_paths(tmp_path, monkeypatch):
    """Redirect import_excel to a temporary DB."""
    db_file = tmp_path / "sales.db"
    monkeypatch.setattr(ie, "DB_PATH", db_file)
    monkeypatch.setattr(db_mod, "DB_PATH", db_file)
    return db_file


def test_import_excel_creates_units_and_purchases(tmp_path, patched_paths):
    wb = _build_workbook([
        {
            "serial": "SN-001", "model": "LT6040",
            "arrival": "2024.06.01", "price": 500000, "lot": "VendorX",
            "fill": YELLOW,  # 在庫
        },
        {
            "serial": "SN-002", "model": "FL5500",
            "arrival": "2024.07.10", "ship": "2024.12.20",
            "buyer": "客先A", "address": "東京都",
            "price": 1200000, "lot": "VendorY",
            # no fill → 出荷済 (default)
        },
        {
            "serial": "SN-003", "model": "LT6040",
            "arrival": "2024.08.01", "price": 480000,
            "fill": GREEN,  # 資産
        },
        {
            # blank serial → skipped
            "model": "LT6040",
        },
    ])
    xlsx = tmp_path / "test.xlsx"
    wb.save(xlsx)

    imported, skipped = ie.import_excel(xlsx)
    assert imported == 3
    assert skipped == 1

    with db_mod.connect(patched_paths) as conn:
        units = list(conn.execute("SELECT serial_no, model, status FROM units ORDER BY id"))
        purchases = list(conn.execute(
            "SELECT u.serial_no, p.amount, p.vendor_name "
            "FROM purchases p JOIN units u ON u.id = p.unit_id "
            "ORDER BY u.id"
        ))
        sales = list(conn.execute(
            "SELECT u.serial_no, s.sale_date, s.customer_name "
            "FROM sales s JOIN units u ON u.id = s.unit_id"
        ))

    assert [(u["serial_no"], u["model"], u["status"]) for u in units] == [
        ("SN-001", "LT6040", "在庫"),
        ("SN-002", "FL5500", "出荷済"),
        ("SN-003", "LT6040", "資産"),
    ]
    assert [(p["serial_no"], p["amount"], p["vendor_name"]) for p in purchases] == [
        ("SN-001", 500000, "VendorX"),
        ("SN-002", 1200000, "VendorY"),
        ("SN-003", 480000, None),
    ]
    # Only the shipped row should have a sales record.
    assert len(sales) == 1
    assert sales[0]["serial_no"] == "SN-002"
    assert sales[0]["sale_date"] == "2024-12-20"
    assert sales[0]["customer_name"] == "客先A"


def test_import_excel_handles_missing_model(tmp_path, patched_paths):
    wb = _build_workbook([
        {"serial": "SN-X", "arrival": "2024.06.01", "fill": YELLOW},
    ])
    xlsx = tmp_path / "test.xlsx"
    wb.save(xlsx)

    imported, skipped = ie.import_excel(xlsx)
    assert imported == 1
    assert skipped == 0

    with db_mod.connect(patched_paths) as conn:
        row = conn.execute("SELECT model FROM units WHERE serial_no=?", ("SN-X",)).fetchone()
    assert row["model"] == "(機種不明)"


def test_import_excel_raises_for_missing_sheet(tmp_path, patched_paths):
    wb = openpyxl.Workbook()
    wb.active.title = "別のシート"
    xlsx = tmp_path / "test.xlsx"
    wb.save(xlsx)

    with pytest.raises(ValueError, match="対象シート"):
        ie.import_excel(xlsx)
